"""Tests de l'adaptateur Excel."""

from __future__ import annotations

import io
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from app.models.constat import NiveauConfiance
from app.models.rapport import Affirmation, Document, Manifeste, Section, Tableau
from app.services.rendu.tableur import rendu_excel


def _document(tableaux: tuple[Tableau, ...] | None = None) -> Document:
    return Document(
        titre="Étude de filière — le cacao",
        sous_titre="Analyse documentée",
        sections=(
            Section(
                titre="Contexte",
                corps="Un paragraphe analytique.",
                affirmations=(
                    Affirmation(
                        texte="La production avoisine 2,2 millions de tonnes.",
                        source="CNRA",
                        date="2025-10-01",
                        methode="rag",
                        confiance=NiveauConfiance.MOYENNE,
                        empreinte="e3b0c44298fc",
                    ),
                ),
            ),
        ),
        tableaux=(
            tableaux
            if tableaux is not None
            else (
                Tableau(
                    titre="Prix",
                    entetes=("Campagne", "Prix"),
                    lignes=(("2025-2026", "1 500"),),
                ),
            )
        ),
        manifeste=Manifeste(
            modele="opencacao-8b",
            version_modele="1.1.0",
            version_app="0.6.75",
            profil_materiel="cpu",
            genere_le=datetime(2026, 7, 29, 10, 0, tzinfo=UTC),
            empreinte_demandeur="a1b2c3d4e5f6",
        ),
    )


def _classeur(octets: bytes):
    return load_workbook(io.BytesIO(octets))


def test_le_rendu_est_un_xlsx_ouvrable():
    assert _classeur(rendu_excel(_document())).sheetnames


def test_la_provenance_a_sa_propre_feuille():
    """Spec §8.4 : c est la que l auditeur voudra trier et filtrer."""
    assert "Provenance" in _classeur(rendu_excel(_document())).sheetnames


def test_le_manifeste_a_sa_propre_feuille():
    assert "Manifeste" in _classeur(rendu_excel(_document())).sheetnames


def test_les_tableaux_de_donnees_ont_leur_feuille():
    classeur = _classeur(rendu_excel(_document()))
    assert "Prix" in classeur.sheetnames
    assert classeur["Prix"].cell(row=1, column=1).value == "Campagne"
    assert classeur["Prix"].cell(row=2, column=2).value == "1 500"


def test_la_feuille_de_provenance_porte_chaque_affirmation():
    feuille = _classeur(rendu_excel(_document()))["Provenance"]
    assert feuille.cell(row=1, column=1).value == "Section"
    assert feuille.cell(row=2, column=3).value == "CNRA"


def test_le_manifeste_est_lisible_en_paires_cle_valeur():
    feuille = _classeur(rendu_excel(_document()))["Manifeste"]
    paires = {
        feuille.cell(row=index, column=1).value: feuille.cell(row=index, column=2).value
        for index in range(1, feuille.max_row + 1)
    }
    assert paires["Modèle"] == "opencacao-8b"
    assert paires["Version applicative"] == "0.6.75"
    assert paires["Empreinte du demandeur"] == "a1b2c3d4e5f6"


def test_un_titre_de_tableau_trop_long_ne_casse_pas_le_classeur():
    """Excel plafonne un nom de feuille a 31 caracteres."""
    long = _document(tableaux=(Tableau(titre="T" * 60, entetes=("A",), lignes=(("x",),)),))
    assert all(len(nom) <= 31 for nom in _classeur(rendu_excel(long)).sheetnames)


def test_un_titre_de_tableau_avec_caracteres_interdits_est_normalise():
    """Excel refuse \\ / * ? : [ ] dans un nom de feuille."""
    sale = _document(tableaux=(Tableau(titre="Prix/2025[A]", entetes=("A",), lignes=(("x",),)),))
    noms = _classeur(rendu_excel(sale)).sheetnames
    assert all(not set(nom) & set("\\/*?:[]") for nom in noms)


def test_deux_tableaux_de_meme_titre_ne_se_percutent_pas():
    doublons = _document(
        tableaux=(
            Tableau(titre="Prix", entetes=("A",), lignes=(("x",),)),
            Tableau(titre="Prix", entetes=("A",), lignes=(("y",),)),
        )
    )
    classeur = _classeur(rendu_excel(doublons))
    feuilles = [nom for nom in classeur.sheetnames if nom.startswith("Prix")]
    assert len(feuilles) == 2


def _xml_des_feuilles(octets: bytes) -> bytes:
    """Concatene le XML brut de toutes les feuilles du classeur.

    On verifie la garantie LA OU elle se joue : dans le fichier livre. Une assertion
    sur l objet openpyxl survivrait a un refactor qui casserait le rendu.
    """
    import zipfile

    with zipfile.ZipFile(io.BytesIO(octets)) as archive:
        return b"".join(
            archive.read(nom) for nom in archive.namelist() if nom.startswith("xl/worksheets/")
        )


_CHARGES_FORMULE = [
    "=1+1",
    "+1",
    "-1",
    "@SUM(A1)",
    "\t=cmd",
    "\r=cmd",
    "\x0b=1+1",
    "\x00=1+1",
    '=HYPERLINK("http://x","c")',
    "=cmd|'/c calc'!A1",
]


@pytest.mark.parametrize("valeur", _CHARGES_FORMULE)
def test_une_valeur_ressemblant_a_une_formule_n_est_jamais_une_formule(valeur):
    """CWE-1236 : openpyxl n echappe rien, et le contenu vient du modele.

    A l ouverture, une cellule commencant par = + - @ est evaluee comme une formule.
    Un livrable transmis a un bailleur ne doit rien executer chez lui. Les charges
    prefixees d un caractere de controle verifient l ORDRE des operations : la purge
    doit preceder le test d amorce, sans quoi la valeur redevient une formule.
    """
    piege = _document(tableaux=(Tableau(titre="T", entetes=("A",), lignes=((valeur,),)),))
    octets = rendu_excel(piege)
    assert b"<f>" not in _xml_des_feuilles(octets)
    assert _classeur(octets)["T"].cell(row=2, column=1).data_type == "s"


@pytest.mark.parametrize("valeur", _CHARGES_FORMULE)
def test_une_formule_dans_une_affirmation_est_neutralisee(valeur):
    """SEUL chemin reellement atteignable : le moteur passe toujours tableaux=()."""
    document = _document()
    piege = Document(
        titre="T",
        sous_titre="S",
        sections=(
            Section(
                titre="Contexte",
                corps="Prose.",
                affirmations=(
                    Affirmation(
                        texte=valeur,
                        source=valeur,
                        date="",
                        methode="rag",
                        confiance=NiveauConfiance.MOYENNE,
                    ),
                ),
            ),
        ),
        tableaux=(),
        manifeste=document.manifeste,
    )
    assert b"<f>" not in _xml_des_feuilles(rendu_excel(piege))


def test_une_formule_dans_un_entete_est_neutralisee():
    piege = _document(tableaux=(Tableau(titre="T", entetes=("=1+1",), lignes=(("x",),)),))
    assert b"<f>" not in _xml_des_feuilles(rendu_excel(piege))


def test_une_valeur_ordinaire_n_est_pas_alteree():
    """La neutralisation ne doit pas defigurer les donnees legitimes."""
    normal = _document(tableaux=(Tableau(titre="T", entetes=("A",), lignes=(("1 500 FCFA",),)),))
    assert _classeur(rendu_excel(normal))["T"].cell(row=2, column=1).value == "1 500 FCFA"


def test_une_valeur_negative_legitime_n_est_pas_defiguree():
    """L apostrophe est une marque de STYLE.

    Ecrite dans le texte, elle arriverait visible chez l auditeur — et toute variation
    de prix negative, ou tout tiret de remplissage, serait abime.
    """
    normal = _document(tableaux=(Tableau(titre="T", entetes=("A",), lignes=(("-120 FCFA",),)),))
    cellule = _classeur(rendu_excel(normal))["T"].cell(row=2, column=1)
    assert cellule.value == "-120 FCFA"
    assert cellule.quotePrefix is True


def test_un_caractere_de_controle_ne_corrompt_pas_le_classeur():
    """openpyxl leve IllegalCharacterError sur un \\x00 non filtre."""
    sale = _document(tableaux=(Tableau(titre="T", entetes=("A",), lignes=(("x\x00y",),)),))
    valeur = _classeur(rendu_excel(sale))["T"].cell(row=2, column=1).value
    assert "\x00" not in str(valeur)


def test_un_document_sans_tableau_garde_l_entete_la_provenance_et_le_manifeste():
    classeur = _classeur(rendu_excel(_document(tableaux=())))
    assert classeur.sheetnames == ["Document", "Provenance", "Manifeste"]


def test_la_mention_d5_figure_dans_le_classeur():
    """« Non contournable » ne peut pas vouloir dire « sauf en Excel » — et le classeur
    est precisement le format que l auditeur ouvrira."""
    document = _document()
    avec_mention = Document(
        titre=document.titre,
        sous_titre=document.sous_titre,
        sections=document.sections,
        tableaux=(),
        manifeste=document.manifeste,
        mention="Document préparatoire. Il ne constitue pas une déclaration de conformité.",
    )
    feuille = _classeur(rendu_excel(avec_mention))["Document"]
    paires = {
        feuille.cell(row=index, column=1).value: feuille.cell(row=index, column=2).value
        for index in range(1, feuille.max_row + 1)
    }
    assert "préparatoire" in paires["Mention"]
    assert paires["Titre"] == document.titre


def test_une_feuille_nommee_provenance_ne_supplante_pas_l_annexe():
    """Sinon la vraie feuille d audit est reléguee en « Provenance1 », en silence."""
    piege = _document(tableaux=(Tableau(titre="Provenance", entetes=("A",), lignes=(("x",),)),))
    classeur = _classeur(rendu_excel(piege))
    assert classeur["Provenance"].cell(row=1, column=1).value == "Section"


def test_un_titre_de_feuille_avec_caractere_de_controle_ne_leve_pas():
    piege = _document(tableaux=(Tableau(titre="Prix\x00 2026", entetes=("A",), lignes=(("x",),)),))
    assert _classeur(rendu_excel(piege)).sheetnames


def test_au_dela_du_dedoublonnage_le_nom_de_repli_prend_la_main():
    """Cent tableaux homonymes : on ne laisse pas openpyxl renommer en silence."""
    homonymes = _document(
        tableaux=tuple(
            Tableau(titre="Prix", entetes=("A",), lignes=((str(index),),)) for index in range(100)
        )
    )
    noms = _classeur(rendu_excel(homonymes)).sheetnames
    assert len(noms) == len(set(noms))  # aucune collision
    assert len(noms) == 103  # 100 tableaux + Document + Provenance + Manifeste


def test_les_metadonnees_du_classeur_sont_celles_du_projet():
    proprietes = _classeur(rendu_excel(_document())).properties
    assert "OpenCacao" in proprietes.creator
    assert "openpyxl" not in proprietes.creator
