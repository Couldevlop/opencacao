"""Rendu Excel — annexes de données et tableau de provenance.

La provenance a sa **feuille dédiée** (spec §8.4) : c'est le format dans lequel un
auditeur voudra la trier et la filtrer.

**Les formules sont neutralisées.** ``openpyxl`` n'échappe rien : une cellule dont la
valeur commence par ``=``, ``+``, ``-`` ou ``@`` est évaluée comme une formule à
l'ouverture du fichier (CWE-1236, « CSV injection »). Le contenu vient du modèle et de
sources externes, et le livrable est transmis à un bailleur ou à un auditeur : il ne
doit rien exécuter chez lui. La garantie vit **ici** et non dans le domaine — préfixer
par une apostrophe est correct pour un tableur et corromprait le Markdown, le Word et
le PowerPoint.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.application.provenance import tableau_de_provenance
from app.models.rapport import Document, Tableau
from app.services.rendu import enrichi
from app.services.rendu.ooxml import texte_xml_sur

# Auteur inscrit dans les métadonnées du fichier livré.
_AUTEUR = "OpenCacao — OpenLab Consulting"

# Excel refuse un nom de feuille de plus de 31 caractères, et les caractères \/*?:[]
_LONGUEUR_MAX_FEUILLE = 31
_INTERDITS_FEUILLE = str.maketrans({caractere: "-" for caractere in "\\/*?:[]"})

# Amorces qui font interpréter une cellule comme une formule. La tabulation et le
# retour chariot comptent aussi : certains tableurs les ignorent avant d'évaluer.
_AMORCES_FORMULE = ("=", "+", "-", "@", "\t", "\r")


def _valeur(brut: str) -> tuple[str, bool]:
    """Rend une valeur inoffensive pour un tableur.

    L'apostrophe de neutralisation n'est **pas** écrite dans le texte : elle est posée
    en marque de style (``quotePrefix``). Écrite dans le texte, elle arriverait visible
    chez l'auditeur — et toute valeur légitime commençant par ``-`` (une variation de
    prix négative, un tiret de remplissage) serait défigurée.

    Args:
        brut: Contenu de la cellule, éventuellement issu du modèle.

    Returns:
        Le couple ``(valeur purgée, faut-il la marquer comme texte)``.
    """
    # Le balisage du modèle est retiré : un tableur ne porte pas d'enrichissement, et
    # « **1 200** » y arriverait tel quel dans la cellule (audit du 19/08).
    propre = enrichi.sans_balisage(texte_xml_sur(str(brut)))
    return propre, propre.startswith(_AMORCES_FORMULE)


def _ajouter_ligne(feuille: Worksheet, valeurs: tuple[str, ...], gras: bool = False) -> None:
    """Ajoute une ligne dont aucune cellule ne peut devenir une formule.

    Args:
        feuille: Feuille de destination.
        valeurs: Contenus des cellules, dans l'ordre des colonnes.
        gras: Met la première cellule en gras (en-tête de tableau ou de paire).
    """
    feuille.append([""] * len(valeurs))
    rangee = feuille.max_row
    for colonne, brut in enumerate(valeurs, start=1):
        valeur, prefixe = _valeur(brut)
        cellule = feuille.cell(row=rangee, column=colonne)
        cellule.value = valeur
        cellule.data_type = "s"
        cellule.quotePrefix = prefixe
        if gras:
            cellule.font = Font(bold=True)


def _nom_de_feuille(titre: str, defaut: str, pris: set[str]) -> str:
    """Rend un titre utilisable et unique comme nom de feuille Excel.

    Args:
        titre: Titre du tableau.
        defaut: Nom de repli si le titre ne donne rien d'utilisable.
        pris: Noms déjà attribués, pour éviter qu'openpyxl ne renomme en silence.

    Returns:
        Un nom de feuille valide, unique, d'au plus 31 caractères.
    """
    propre = texte_xml_sur(titre or defaut).translate(_INTERDITS_FEUILLE).strip()
    nom = (propre or defaut)[:_LONGUEUR_MAX_FEUILLE]
    if nom not in pris:
        return nom
    for suffixe in range(2, 100):
        candidat = f"{nom[: _LONGUEUR_MAX_FEUILLE - len(str(suffixe)) - 1]} {suffixe}"
        if candidat not in pris:
            return candidat
    return defaut[:_LONGUEUR_MAX_FEUILLE]


def _ecrire_tableau(feuille: Worksheet, tableau: Tableau) -> None:
    """Écrit un tableau (en-têtes en gras) dans une feuille.

    Args:
        feuille: Feuille de destination.
        tableau: Tableau à écrire.
    """
    _ajouter_ligne(feuille, tableau.entetes, gras=True)
    for ligne in tableau.lignes:
        _ajouter_ligne(feuille, ligne)


def rendu_excel(document: Document) -> bytes:
    """Rend le document au format Excel.

    Args:
        document: Document assemblé par le moteur.

    Returns:
        Les octets du fichier ``.xlsx``.
    """
    classeur = Workbook()
    classeur.remove(classeur.active)
    classeur.properties.creator = _AUTEUR
    classeur.properties.lastModifiedBy = _AUTEUR
    classeur.properties.title = texte_xml_sur(document.titre)
    classeur.properties.created = document.manifeste.genere_le
    classeur.properties.modified = document.manifeste.genere_le

    # D5 EN PREMIER. « Non contournable » ne peut pas vouloir dire « sauf en Excel » :
    # le classeur est précisément le format que l'auditeur ouvrira. Sans cette feuille,
    # le dossier de parcelle partait sans sa mention préparatoire — ni même son titre.
    entete = classeur.create_sheet("Document")
    for cle, valeur in (
        ("Titre", document.titre),
        ("Sous-titre", document.sous_titre),
        ("Mention", document.mention),
    ):
        _ajouter_ligne(entete, (cle, valeur), gras=True)
    # Réservés dès le départ : sinon un tableau intitulé « Provenance » prend le nom
    # et la vraie feuille d'audit est reléguée en « Provenance1 », en silence.
    pris: set[str] = {"Document", "Provenance", "Manifeste"}

    for index, tableau in enumerate(document.tableaux, start=1):
        nom = _nom_de_feuille(tableau.titre, f"Tableau {index}", pris)
        pris.add(nom)
        _ecrire_tableau(classeur.create_sheet(nom), tableau)

    _ecrire_tableau(classeur.create_sheet("Provenance"), tableau_de_provenance(document))

    manifeste = document.manifeste
    feuille = classeur.create_sheet("Manifeste")
    for cle, valeur in (
        ("Modèle", manifeste.modele),
        ("Version du modèle", manifeste.version_modele),
        ("Version applicative", manifeste.version_app),
        ("Profil matériel", manifeste.profil_materiel),
        ("Généré le", manifeste.genere_le.isoformat()),
        ("Empreinte du demandeur", manifeste.empreinte_demandeur),
    ):
        _ajouter_ligne(feuille, (cle, valeur), gras=True)

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
