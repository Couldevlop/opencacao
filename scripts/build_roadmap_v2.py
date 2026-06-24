"""Génère la roadmap agile OpenCacao V2 (conversationnel) au format Excel.

Usage : python scripts/build_roadmap_v2.py
Sortie : docs/OpenCacao_V2_Roadmap_Agile.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# --------------------------------------------------------------------------
# Palette (cohérente avec la marque OpenLab Consulting : orange + anthracite)
# --------------------------------------------------------------------------
ORANGE = "EA5B13"
ORANGE_DARK = "B8410A"
ORANGE_LIGHT = "FCE7D8"
ORANGE_PALE = "FDF4EE"
DARK = "1F1F1F"
GREY_HEAD = "3F3F3F"
GREY_LIGHT = "F2F2F2"
WHITE = "FFFFFF"

MOSCOW = {
    "Must": ("E53935", WHITE),
    "Should": ("FB8C00", WHITE),
    "Could": ("FDD835", DARK),
    "Won't": ("9E9E9E", WHITE),
}
STATUT = {
    "À faire": ("E4E4E4", DARK),
    "En cours": ("BBDEFB", DARK),
    "Fait": ("C8E6C9", DARK),
}
SPRINT_COLOR = {
    "Sprint 1": "EA5B13",
    "Sprint 2": "F3833E",
    "Sprint 3": "1565C0",
    "Sprint 4": "2E7D32",
    "Sprint 5": "6A1B9A",
    "Sprint 6": "00838F",
}

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROOT = Path(__file__).resolve().parent.parent
LOGO = ROOT / "docs" / "OPENLAB.png"
OUT = ROOT / "docs" / "OpenCacao_V2_Roadmap_Agile.xlsx"


def fill(hexcolor: str) -> PatternFill:
    return PatternFill("solid", fgColor=hexcolor)


def header_row(ws, row, headers, *, bg=ORANGE, fg=WHITE, height=26):
    ws.row_dimensions[row].height = height
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill = fill(bg)
        cell.font = Font(bold=True, color=fg, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, subtitle=None, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1, value=title)
    t.fill = fill(DARK)
    t.font = Font(bold=True, color=WHITE, size=16)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.fill = fill(ORANGE)
        s.font = Font(bold=False, color=WHITE, size=10, italic=True)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20


wb = Workbook()

# ==========================================================================
# 1. PAGE DE GARDE
# ==========================================================================
cover = wb.active
cover.title = "Page de garde"
cover.sheet_view.showGridLines = False
for col in "ABCDEFGHIJ":
    cover.column_dimensions[col].width = 13

# bande supérieure orange
cover.merge_cells("A1:J1")
cover["A1"].fill = fill(ORANGE)
cover.row_dimensions[1].height = 10

# logo centré
with PILImage.open(LOGO) as im:
    w, h = im.size
target_w = 360
ratio = target_w / w
img = XLImage(str(LOGO))
img.width = target_w
img.height = int(h * ratio)
img.anchor = "D3"
cover.add_image(img)
for r in range(2, 13):
    cover.row_dimensions[r].height = 22

# titre projet
cover.merge_cells("A15:J15")
cover["A15"] = "OpenCacao — Version 2"
cover["A15"].font = Font(bold=True, size=28, color=DARK)
cover["A15"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[15].height = 44

cover.merge_cells("A16:J16")
cover["A16"] = "Assistant 100 % conversationnel — mémoire & sessions persistantes"
cover["A16"].font = Font(size=14, color=ORANGE_DARK, italic=True)
cover["A16"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[16].height = 26

cover.merge_cells("A17:J17")
cover["A17"] = "Feuille de route — Gestion de projet Agile (Scrum)"
cover["A17"].font = Font(size=12, color=GREY_HEAD)
cover["A17"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[17].height = 22

# filet
cover.merge_cells("C19:H19")
cover["C19"].fill = fill(ORANGE)
cover.row_dimensions[19].height = 4

# bloc métadonnées
meta = [
    ("Maître d'ouvrage", "OpenLab Consulting"),
    ("Sponsor / Validation", "Waopron Coulibaly"),
    ("Produit", "OpenCacao-8B — conseil agronomique cacao (Côte d'Ivoire)"),
    ("Périmètre", "V2 conversationnelle : sessions, mémoire durable, interface type ChatGPT/Claude"),
    ("Méthodologie", "Scrum — 6 sprints de 2 semaines"),
    ("Date de lancement", "30 juin 2026"),
    ("Mise en production cible", "19 septembre 2026 (≈ 12 semaines)"),
    ("Version du document", "1.0 — 24 juin 2026"),
    ("Confidentialité", "Interne — démonstration souveraine (livre blanc IA pour la Côte d'Ivoire)"),
]
start = 21
for i, (k, v) in enumerate(meta):
    r = start + i
    cover.row_dimensions[r].height = 22
    cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    kc = cover.cell(row=r, column=2, value=k)
    kc.fill = fill(DARK)
    kc.font = Font(bold=True, color=WHITE, size=10)
    kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    vc = cover.cell(row=r, column=4, value=v)
    vc.fill = fill(ORANGE_PALE if i % 2 == 0 else WHITE)
    vc.font = Font(color=DARK, size=10)
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# bande inférieure
foot = start + len(meta) + 1
cover.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=10)
fc = cover.cell(row=foot, column=1, value="OpenLab Consulting — IA souveraine pour la Côte d'Ivoire")
fc.fill = fill(ORANGE)
fc.font = Font(bold=True, color=WHITE, size=10)
fc.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[foot].height = 22

# ==========================================================================
# 2. ÉTAT DES LIEUX
# ==========================================================================
etat = wb.create_sheet("1. État des lieux")
etat.sheet_view.showGridLines = False
set_widths(etat, {"A": 3, "B": 34, "C": 16, "D": 70})
title_block(etat, "Où en sommes-nous ?", "Audit du code au 24 juin 2026 — base de départ de la V2", span=4)

header_row(etat, 4, ["", "Composant", "Statut", "Détail"], bg=GREY_HEAD)
rows = [
    ("Multi-tours (base)", "Existant", "Dialogue multi-tours géré : historique envoyé par le client à chaque requête."),
    ("Clarification consultative", "Existant", "Questions complémentaires sur symptômes vagues avant de répondre."),
    ("Ré-ancrage RAG", "Existant", "Requête RAG ré-ancrée sur le thème en cours (anti-dérive multi-tours)."),
    ("Streaming SSE", "Existant", "Réponse en flux (/v1/chat/stream)."),
    ("Garde-fous métier", "Existant", "Refus ANADER, évalués tour par tour (sans mémoire des tours passés)."),
    ("Sessions de conversation", "Absent", "Aucun modèle, aucun endpoint, aucune notion de session côté serveur."),
    ("Persistance de l'historique", "Absent", "Historique volatile en mémoire navigateur (perdu au rechargement, max 20 msg)."),
    ("Base de données", "Absent", "Aucune BD durable. Redis = cache + rate-limit uniquement."),
    ("Mémoire serveur", "Absent", "Serveur sans état : ne reconstruit pas le contexte."),
    ("Sidebar / liste de conversations", "Absent", "UI mono-conversation, pas de navigation ni de reprise."),
    ("Titre auto de conversation", "Absent", "Aucune génération de titre."),
    ("Recherche / reprise / renommage", "Absent", "Aucun moyen de retrouver ou rouvrir une conversation passée."),
    ("Identité utilisateur", "Absent", "Pas de compte ni d'identifiant persistant."),
]
r = 5
for comp, st, det_txt in rows:
    zebra = ORANGE_PALE if (r % 2 == 1) else WHITE
    cc = etat.cell(row=r, column=2, value=comp)
    cc.font = Font(bold=True, size=10, color=DARK)
    cc.fill = fill(zebra)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc = etat.cell(row=r, column=3, value=st)
    sc.fill = fill("C8E6C9" if st == "Existant" else "FFCDD2")
    sc.font = Font(bold=True, size=10, color=DARK)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    dc = etat.cell(row=r, column=4, value=det_txt)
    dc.font = Font(size=10, color=DARK)
    dc.fill = fill(zebra)
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in (2, 3, 4):
        etat.cell(row=r, column=c).border = BORDER
    etat.row_dimensions[r].height = 30
    r += 1

concl = r + 1
etat.merge_cells(start_row=concl, start_column=2, end_row=concl, end_column=4)
cc = etat.cell(row=concl, column=2,
               value="Conclusion : le socle dialogue existe, mais 100 % de la couche « sessions + mémoire persistante + UI conversationnelle » reste à construire. C'est l'objet de cette V2.")
cc.fill = fill(ORANGE_LIGHT)
cc.font = Font(bold=True, size=10, color=ORANGE_DARK)
cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
etat.row_dimensions[concl].height = 40

# ==========================================================================
# 3. BACKLOG PRODUIT
# ==========================================================================
backlog = wb.create_sheet("2. Backlog produit")
backlog.sheet_view.showGridLines = False
set_widths(backlog, {"A": 8, "B": 26, "C": 54, "D": 11, "E": 9, "F": 11})
title_block(backlog, "Backlog produit — Epics & User Stories",
            "Priorisation MoSCoW · estimation en Story Points (SP)", span=6)
header_row(backlog, 4, ["ID", "Epic", "User Story", "Priorité", "SP", "Sprint"])

backlog_rows = [
    ("A — Persistance & Sessions (Backend)", None, None, None, None),
    ("A1", "A — Persistance & Sessions", "Définir les modèles Session & ConversationMessage (Pydantic v2).", "Must", 5, "Sprint 1"),
    ("A2", "A — Persistance & Sessions", "Couche de stockage durable (SQLite dev / PostgreSQL prod) + migrations.", "Must", 8, "Sprint 1"),
    ("A3", "A — Persistance & Sessions", "Endpoints CRUD /v1/sessions (créer, lister, obtenir, supprimer).", "Must", 5, "Sprint 2"),
    ("A4", "A — Persistance & Sessions", "Ajouter session_id à /v1/chat et /v1/chat/stream.", "Must", 3, "Sprint 2"),
    ("A5", "A — Persistance & Sessions", "Persister l'historique côté serveur à chaque tour.", "Must", 5, "Sprint 2"),
    ("B — Mémoire conversationnelle", None, None, None, None),
    ("B1", "B — Mémoire conversationnelle", "Reconstruire le contexte côté serveur depuis la session.", "Must", 5, "Sprint 3"),
    ("B2", "B — Mémoire conversationnelle", "Fenêtre glissante + résumé automatique au-delà de N tours.", "Should", 8, "Sprint 3"),
    ("B3", "B — Mémoire conversationnelle", "Génération automatique du titre de conversation.", "Should", 5, "Sprint 3"),
    ("B4", "B — Mémoire conversationnelle", "Ré-ancrer garde-fous & RAG sur le fil complet.", "Should", 3, "Sprint 3"),
    ("C — Interface conversationnelle (Frontend)", None, None, None, None),
    ("C1", "C — Interface conversationnelle", "Refonte du layout avec sidebar responsive.", "Must", 8, "Sprint 4"),
    ("C2", "C — Interface conversationnelle", "Liste des conversations : création / sélection / reprise.", "Must", 5, "Sprint 4"),
    ("C4", "C — Interface conversationnelle", "Persistance locale des IDs + synchronisation API.", "Must", 3, "Sprint 4"),
    ("C3", "C — Interface conversationnelle", "Renommer & supprimer une conversation.", "Should", 3, "Sprint 5"),
    ("C5", "C — Interface conversationnelle", "Recherche plein-texte dans les conversations.", "Could", 5, "Sprint 5"),
    ("D — Identité & continuité", None, None, None, None),
    ("D1", "D — Identité & continuité", "Identifiant anonyme persistant (device / cookie).", "Should", 3, "Sprint 5"),
    ("D2", "D — Identité & continuité", "Authentification légère optionnelle (magic link / code).", "Could", 8, "Sprint 5"),
    ("E — Qualité, sécurité & déploiement", None, None, None, None),
    ("E1", "E — Qualité & déploiement", "Tests unitaires + intégration (≥ 80 % sur les nouveaux modules).", "Must", 5, "Sprint 6"),
    ("E2", "E — Qualité & déploiement", "RGPD : rétention, purge auto, export, anonymisation.", "Must", 5, "Sprint 6"),
    ("E3", "E — Qualité & déploiement", "Déploiement prod K3s : volume persistant PostgreSQL + migration.", "Must", 5, "Sprint 6"),
    ("E4", "E — Qualité & déploiement", "Documentation + mise à jour CLAUDE_OpenCacao.md (spec V2).", "Should", 3, "Sprint 6"),
    ("E5", "E — Qualité & déploiement", "Recette utilisateur & validation finale (Waopron Coulibaly).", "Must", 3, "Sprint 6"),
]
r = 5
for row in backlog_rows:
    if row[1] is None:  # ligne de groupe Epic
        backlog.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        gc = backlog.cell(row=r, column=1, value=row[0])
        gc.fill = fill(DARK)
        gc.font = Font(bold=True, color=WHITE, size=11)
        gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        backlog.row_dimensions[r].height = 22
        r += 1
        continue
    rid, epic, story, prio, sp, sprint = row
    backlog.cell(row=r, column=1, value=rid).font = Font(bold=True, size=10)
    backlog.cell(row=r, column=2, value=epic.split("—")[0].strip()).font = Font(size=9, color=GREY_HEAD)
    backlog.cell(row=r, column=3, value=story).font = Font(size=10, color=DARK)
    pc = backlog.cell(row=r, column=4, value=prio)
    pbg, pfg = MOSCOW[prio]
    pc.fill = fill(pbg)
    pc.font = Font(bold=True, color=pfg, size=10)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    spc = backlog.cell(row=r, column=5, value=sp)
    spc.alignment = Alignment(horizontal="center", vertical="center")
    spc.font = Font(bold=True, size=10)
    sc = backlog.cell(row=r, column=6, value=sprint)
    sc.fill = fill(SPRINT_COLOR[sprint])
    sc.font = Font(bold=True, color=WHITE, size=9)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    backlog.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    backlog.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    backlog.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in range(1, 7):
        backlog.cell(row=r, column=c).border = BORDER
    backlog.row_dimensions[r].height = 26
    r += 1

total = r
backlog.merge_cells(start_row=total, start_column=1, end_row=total, end_column=4)
backlog.cell(row=total, column=1, value="TOTAL").font = Font(bold=True, size=11, color=WHITE)
backlog.cell(row=total, column=1).fill = fill(ORANGE)
backlog.cell(row=total, column=1).alignment = Alignment(horizontal="right", vertical="center")
total_sp = sum(x[4] for x in backlog_rows if x[1] is not None)
tc = backlog.cell(row=total, column=5, value=total_sp)
tc.fill = fill(ORANGE)
tc.font = Font(bold=True, size=11, color=WHITE)
tc.alignment = Alignment(horizontal="center", vertical="center")
backlog.cell(row=total, column=6, value="≈17/sprint").fill = fill(ORANGE)
backlog.cell(row=total, column=6).font = Font(bold=True, size=9, color=WHITE)
backlog.cell(row=total, column=6).alignment = Alignment(horizontal="center", vertical="center")

# ==========================================================================
# 4. PLANNING DES SPRINTS
# ==========================================================================
spr = wb.create_sheet("3. Planning des sprints")
spr.sheet_view.showGridLines = False
set_widths(spr, {"A": 11, "B": 24, "C": 60, "D": 14, "E": 12})
title_block(spr, "Planning des sprints", "Scrum — itérations de 2 semaines · cérémonies à chaque sprint", span=5)
header_row(spr, 4, ["Sprint", "Période", "Objectif (incrément livrable)", "User Stories", "Capacité (SP)"])

sprints = [
    ("Sprint 1", "30/06 → 11/07", "Fondations : modèles de données + stockage durable + migrations.", "A1, A2", 13),
    ("Sprint 2", "14/07 → 25/07", "API de sessions : CRUD + chat persistant côté serveur.", "A3, A4, A5", 13),
    ("Sprint 3", "28/07 → 08/08", "Mémoire conversationnelle : contexte serveur, résumé, titres auto.", "B1, B2, B3, B4", 21),
    ("Sprint 4", "11/08 → 22/08", "MVP UI : sidebar + liste de conversations + reprise de session.", "C1, C2, C4", 16),
    ("Sprint 5", "25/08 → 05/09", "UI avancée + identité : renommage, recherche, identifiant, auth (option).", "C3, C5, D1, D2", 19),
    ("Sprint 6", "08/09 → 19/09", "Durcissement : tests, RGPD, déploiement prod, doc, recette finale.", "E1, E2, E3, E4, E5", 21),
]
r = 5
for s, per, obj, us, cap in sprints:
    sc = spr.cell(row=r, column=1, value=s)
    sc.fill = fill(SPRINT_COLOR[s])
    sc.font = Font(bold=True, color=WHITE, size=11)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=2, value=per).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=2).font = Font(bold=True, size=10)
    spr.cell(row=r, column=3, value=obj).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    spr.cell(row=r, column=3).font = Font(size=10)
    spr.cell(row=r, column=4, value=us).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=4).font = Font(size=10)
    spr.cell(row=r, column=5, value=cap).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=5).font = Font(bold=True, size=10)
    for c in range(1, 6):
        spr.cell(row=r, column=c).border = BORDER
    spr.row_dimensions[r].height = 38
    r += 1

note_r = r + 1
spr.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=5)
nc = spr.cell(row=note_r, column=1,
              value="Cérémonies par sprint : Sprint Planning (J1) · Daily Scrum (15 min/j) · Sprint Review + démo (dernier jour) · Rétrospective. Définition de « Terminé » : code revu, testé (≥80 %), garde-fous verts, documenté, déployé en recette.")
nc.fill = fill(ORANGE_PALE)
nc.font = Font(italic=True, size=10, color=DARK)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
spr.row_dimensions[note_r].height = 44

# ==========================================================================
# 5. TÂCHES DÉTAILLÉES
# ==========================================================================
det = wb.create_sheet("4. Tâches détaillées")
det.sheet_view.showGridLines = False
set_widths(det, {"A": 7, "B": 56, "C": 16, "D": 9, "E": 11, "F": 11, "G": 11, "H": 11})
title_block(det, "Tâches détaillées (planning prévisionnel)",
            "Estimation en jours-homme (j-h) · dates ouvrées 2026", span=8)
header_row(det, 4, ["ID", "Tâche", "Rôle", "j-h", "Début", "Fin", "Statut", "Sprint"])

tasks = [
    # Sprint 1
    ("T1.1", "Cadrage technique V2 + choix BD (SQLite dev / PostgreSQL prod)", "Lead", 1.5, "30/06", "01/07", "Sprint 1"),
    ("T1.2", "Modèles Pydantic Session & ConversationMessage", "Dev Backend", 1.5, "01/07", "02/07", "Sprint 1"),
    ("T1.3", "Schéma BD + migrations (Alembic)", "Dev Backend", 2, "03/07", "06/07", "Sprint 1"),
    ("T1.4", "Repository session (CRUD bas niveau) + tests", "Dev Backend", 3, "07/07", "10/07", "Sprint 1"),
    ("T1.5", "Config BD (core/config.py) + service db (docker-compose)", "DevOps", 1, "10/07", "11/07", "Sprint 1"),
    # Sprint 2
    ("T2.1", "Endpoints CRUD /v1/sessions (créer/lister/obtenir/supprimer)", "Dev Backend", 3, "14/07", "16/07", "Sprint 2"),
    ("T2.2", "Ajout du champ session_id à ChatRequest + validation", "Dev Backend", 1, "16/07", "17/07", "Sprint 2"),
    ("T2.3", "Chargement/sauvegarde de l'historique dans /v1/chat", "Dev Backend", 3, "17/07", "22/07", "Sprint 2"),
    ("T2.4", "Idem pour /v1/chat/stream (SSE)", "Dev Backend", 2, "22/07", "24/07", "Sprint 2"),
    ("T2.5", "Tests d'intégration des endpoints sessions (≥80 %)", "QA", 2, "23/07", "25/07", "Sprint 2"),
    # Sprint 3
    ("T3.1", "Reconstruction du contexte côté serveur depuis la session", "Dev Backend", 2, "28/07", "29/07", "Sprint 3"),
    ("T3.2", "Fenêtre glissante + troncature intelligente du contexte", "Dev Backend", 2, "30/07", "31/07", "Sprint 3"),
    ("T3.3", "Résumé automatique des tours anciens (via le modèle)", "Dev Backend", 3, "31/07", "05/08", "Sprint 3"),
    ("T3.4", "Génération automatique du titre de conversation", "Dev Backend", 2, "05/08", "07/08", "Sprint 3"),
    ("T3.5", "Ré-ancrage garde-fous + RAG sur le fil complet + tests", "Dev BE / QA", 2, "06/08", "08/08", "Sprint 3"),
    # Sprint 4
    ("T4.1", "Maquette UI sidebar + validation PO", "PO / UX", 1.5, "11/08", "12/08", "Sprint 4"),
    ("T4.2", "Refonte layout + sidebar responsive (HTML/CSS)", "Dev Frontend", 3, "12/08", "15/08", "Sprint 4"),
    ("T4.3", "Client API sessions (sessions-api-client.js)", "Dev Frontend", 2, "15/08", "18/08", "Sprint 4"),
    ("T4.4", "Liste conversations : création / sélection / reprise", "Dev Frontend", 3, "18/08", "21/08", "Sprint 4"),
    ("T4.5", "Persistance locale des IDs + synchronisation API", "Dev Frontend", 1.5, "21/08", "22/08", "Sprint 4"),
    # Sprint 5
    ("T5.1", "Renommer & supprimer une conversation (UI + API)", "Dev FE / BE", 2, "25/08", "27/08", "Sprint 5"),
    ("T5.2", "Recherche plein-texte dans les conversations", "Dev BE / FE", 3, "27/08", "01/09", "Sprint 5"),
    ("T5.3", "Identifiant anonyme persistant (device id / cookie)", "Dev Backend", 2, "01/09", "03/09", "Sprint 5"),
    ("T5.4", "Authentification légère optionnelle (magic link) — Could", "Dev BE / FE", 3, "03/09", "05/09", "Sprint 5"),
    # Sprint 6
    ("T6.1", "Durcissement des tests (couverture ≥ 80 % globale)", "QA", 3, "08/09", "10/09", "Sprint 6"),
    ("T6.2", "RGPD : rétention, purge auto, export, anonymisation", "Dev Backend", 3, "10/09", "12/09", "Sprint 6"),
    ("T6.3", "Déploiement prod K3s : PVC PostgreSQL + migration", "DevOps", 2, "15/09", "16/09", "Sprint 6"),
    ("T6.4", "Documentation + mise à jour CLAUDE_OpenCacao.md (spec V2)", "Lead", 2, "16/09", "18/09", "Sprint 6"),
    ("T6.5", "Recette utilisateur & validation finale (Waopron Coulibaly)", "PO", 1, "19/09", "19/09", "Sprint 6"),
]
r = 5
for tid, name, role, jh, deb, fin, sprint in tasks:
    det.cell(row=r, column=1, value=tid).font = Font(bold=True, size=9)
    det.cell(row=r, column=2, value=name).font = Font(size=10)
    det.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    det.cell(row=r, column=3, value=role).font = Font(size=9, color=GREY_HEAD)
    det.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
    det.cell(row=r, column=4, value=jh).alignment = Alignment(horizontal="center", vertical="center")
    det.cell(row=r, column=5, value=deb).alignment = Alignment(horizontal="center", vertical="center")
    det.cell(row=r, column=6, value=fin).alignment = Alignment(horizontal="center", vertical="center")
    sc = det.cell(row=r, column=7, value="À faire")
    sbg, sfg = STATUT["À faire"]
    sc.fill = fill(sbg)
    sc.font = Font(size=9, color=sfg)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    spc = det.cell(row=r, column=8, value=sprint)
    spc.fill = fill(SPRINT_COLOR[sprint])
    spc.font = Font(bold=True, size=8, color=WHITE)
    spc.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 9):
        det.cell(row=r, column=c).border = BORDER
        if c in (1, 5, 6):
            det.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    det.row_dimensions[r].height = 28
    r += 1

det.freeze_panes = "A5"

# ==========================================================================
# 6. RÉTRO-PLANNING & JALONS
# ==========================================================================
retro = wb.create_sheet("5. Rétro-planning")
retro.sheet_view.showGridLines = False
set_widths(retro, {"A": 8, "B": 40, "C": 18, "D": 16, "E": 44})
title_block(retro, "Rétro-planning — jalons à rebours de la cible",
            "Date de mise en production : 19 septembre 2026 · on remonte du livrable final vers le lancement", span=5)
header_row(retro, 4, ["Jalon", "Livrable / Definition of Done", "Échéance au + tard", "Marge", "Pré-requis"])

jalons = [
    ("M6", "Recette validée & mise en production V2 (Go-Live).", "19/09/2026", "0 j", "M5 + déploiement + RGPD OK"),
    ("M5", "Feature complete : recherche, identité, options livrées.", "05/09/2026", "2 sem.", "M4 + UI avancée"),
    ("M4", "MVP UI conversationnelle : sidebar + reprise de session.", "22/08/2026", "2 sem.", "M3 + API stable"),
    ("M3", "Mémoire conversationnelle complète (contexte serveur, titres).", "08/08/2026", "2 sem.", "M2 + persistance"),
    ("M2", "API de sessions opérationnelle (CRUD + chat persistant).", "25/07/2026", "2 sem.", "M1 + endpoints"),
    ("M1", "Fondations de persistance (BD + repository + migrations).", "11/07/2026", "2 sem.", "M0 + cadrage validé"),
    ("M0", "Lancement projet & cadrage technique validé.", "30/06/2026", "—", "Backlog priorisé, équipe mobilisée"),
]
r = 5
for jid, livr, ech, marge, pre in jalons:
    jc = retro.cell(row=r, column=1, value=jid)
    jc.fill = fill(ORANGE if jid == "M6" else DARK)
    jc.font = Font(bold=True, color=WHITE, size=11)
    jc.alignment = Alignment(horizontal="center", vertical="center")
    retro.cell(row=r, column=2, value=livr).font = Font(size=10)
    retro.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ec = retro.cell(row=r, column=3, value=ech)
    ec.font = Font(bold=True, size=10, color=ORANGE_DARK)
    ec.alignment = Alignment(horizontal="center", vertical="center")
    retro.cell(row=r, column=4, value=marge).alignment = Alignment(horizontal="center", vertical="center")
    retro.cell(row=r, column=5, value=pre).font = Font(size=9, color=GREY_HEAD)
    retro.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in range(1, 6):
        retro.cell(row=r, column=c).border = BORDER
    retro.row_dimensions[r].height = 34
    r += 1

cr = r + 1
retro.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=5)
crit = retro.cell(row=cr, column=1,
                  value="Chemin critique : M0 → M1 → M2 → M3 → M4 → M6. La recherche, l'auth (D2) et le résumé (B2) sont des éléments « Should/Could » qui peuvent glisser sans décaler le Go-Live.")
crit.fill = fill(ORANGE_LIGHT)
crit.font = Font(italic=True, size=10, color=ORANGE_DARK)
crit.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
retro.row_dimensions[cr].height = 36

# ==========================================================================
# 7. DIAGRAMME DE GANTT
# ==========================================================================
gantt = wb.create_sheet("6. Gantt")
gantt.sheet_view.showGridLines = False
weeks = ["S1\n30/06", "S2\n07/07", "S3\n14/07", "S4\n21/07", "S5\n28/07", "S6\n04/08",
         "S7\n11/08", "S8\n18/08", "S9\n25/08", "S10\n01/09", "S11\n08/09", "S12\n15/09"]
set_widths(gantt, {"A": 8, "B": 42})
for i in range(len(weeks)):
    gantt.column_dimensions[get_column_letter(3 + i)].width = 7
title_block(gantt, "Diagramme de Gantt — vue hebdomadaire", "12 semaines · 6 sprints de 2 semaines", span=2 + len(weeks))

# en-têtes semaines
gantt.row_dimensions[4].height = 30
gantt.cell(row=4, column=1, value="ID").fill = fill(GREY_HEAD)
gantt.cell(row=4, column=2, value="User Story").fill = fill(GREY_HEAD)
for c in (1, 2):
    gantt.cell(row=4, column=c).font = Font(bold=True, color=WHITE, size=10)
    gantt.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center")
    gantt.cell(row=4, column=c).border = BORDER
for i, wlabel in enumerate(weeks):
    cell = gantt.cell(row=4, column=3 + i, value=wlabel)
    cell.fill = fill(GREY_HEAD)
    cell.font = Font(bold=True, color=WHITE, size=8)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

# sprint -> plage de semaines (0-indexées)
sprint_weeks = {
    "Sprint 1": (0, 1), "Sprint 2": (2, 3), "Sprint 3": (4, 5),
    "Sprint 4": (6, 7), "Sprint 5": (8, 9), "Sprint 6": (10, 11),
}
gantt_rows = [
    ("A1", "Modèles Session & Message", "Sprint 1"),
    ("A2", "Stockage durable + migrations", "Sprint 1"),
    ("A3", "Endpoints CRUD /v1/sessions", "Sprint 2"),
    ("A4", "session_id dans /v1/chat", "Sprint 2"),
    ("A5", "Historique persistant serveur", "Sprint 2"),
    ("B1", "Contexte reconstruit serveur", "Sprint 3"),
    ("B2", "Fenêtre glissante + résumé", "Sprint 3"),
    ("B3", "Titre auto de conversation", "Sprint 3"),
    ("B4", "Garde-fous & RAG sur le fil", "Sprint 3"),
    ("C1", "Layout + sidebar responsive", "Sprint 4"),
    ("C2", "Liste conversations + reprise", "Sprint 4"),
    ("C4", "Persistance locale + synchro", "Sprint 4"),
    ("C3", "Renommer / supprimer", "Sprint 5"),
    ("C5", "Recherche plein-texte", "Sprint 5"),
    ("D1", "Identifiant anonyme", "Sprint 5"),
    ("D2", "Auth légère (option)", "Sprint 5"),
    ("E1", "Tests ≥ 80 %", "Sprint 6"),
    ("E2", "RGPD : rétention / purge", "Sprint 6"),
    ("E3", "Déploiement prod K3s", "Sprint 6"),
    ("E4", "Documentation + spec V2", "Sprint 6"),
    ("E5", "Recette & validation finale", "Sprint 6"),
]
r = 5
for rid, label, sprint in gantt_rows:
    gantt.cell(row=r, column=1, value=rid).font = Font(bold=True, size=9)
    gantt.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    gantt.cell(row=r, column=2, value=label).font = Font(size=9)
    gantt.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    gantt.cell(row=r, column=1).border = BORDER
    gantt.cell(row=r, column=2).border = BORDER
    w0, w1 = sprint_weeks[sprint]
    for i in range(len(weeks)):
        cell = gantt.cell(row=r, column=3 + i)
        cell.border = BORDER
        if w0 <= i <= w1:
            cell.fill = fill(SPRINT_COLOR[sprint])
    gantt.row_dimensions[r].height = 18
    r += 1

# légende jalons
leg = r + 1
gantt.cell(row=leg, column=1, value="Jalons :").font = Font(bold=True, size=9)
milestones = {"Sprint 1": "M1", "Sprint 2": "M2", "Sprint 3": "M3", "Sprint 4": "M4", "Sprint 5": "M5", "Sprint 6": "M6"}
for s, (w0, w1) in sprint_weeks.items():
    cell = gantt.cell(row=leg, column=3 + w1, value=milestones[s])
    cell.fill = fill(ORANGE)
    cell.font = Font(bold=True, color=WHITE, size=8)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
gantt.freeze_panes = "C5"

# ==========================================================================
# 8. RISQUES
# ==========================================================================
risk = wb.create_sheet("7. Risques")
risk.sheet_view.showGridLines = False
set_widths(risk, {"A": 6, "B": 46, "C": 12, "D": 12, "E": 56})
title_block(risk, "Registre des risques", "Suivi en rétrospective de sprint", span=5)
header_row(risk, 4, ["ID", "Risque", "Impact", "Probabilité", "Mitigation"])

def lvl_fill(v):
    return {"Élevé": "FFCDD2", "Moyen": "FFE0B2", "Faible": "C8E6C9",
            "Élevée": "FFCDD2", "Moyenne": "FFE0B2", "Faible ": "C8E6C9"}.get(v, "FFE0B2")

risks = [
    ("R1", "Latence du multi-tours / fenêtre de contexte trop large sur CPU (GGUF llama.cpp, nœud CX53).", "Élevé", "Moyenne", "Résumé automatique + fenêtre glissante + plafond de tokens ; mesurer la latence par sprint."),
    ("R2", "Capacité du serveur Hetzner CX53 insuffisante pour BD + inférence simultanées.", "Élevé", "Moyenne", "SQLite léger d'abord, sizing avant PostgreSQL, monitoring mémoire/CPU, volume dédié."),
    ("R3", "Dérive de périmètre (authentification complète, recherche avancée).", "Moyen", "Élevée", "Priorisation MoSCoW stricte ; D2 et C5 restent « Could » et peuvent glisser."),
    ("R4", "Conformité RGPD : conservation des conversations des producteurs.", "Élevé", "Faible", "Politique de rétention + purge automatique + consentement + anonymisation dès la conception."),
    ("R5", "Régression des garde-fous métier ANADER en contexte multi-tours.", "Élevé", "Moyenne", "Tests garde-fous par tour + jeu d'évaluation figé rejoué à chaque sprint."),
    ("R6", "Disponibilité réduite de l'équipe (petite structure).", "Moyen", "Moyenne", "Sprints avec marge, chemin critique protégé, éléments optionnels en fin de parcours."),
    ("R7", "Perte de données de session lors d'une migration BD en prod.", "Élevé", "Faible", "Sauvegardes avant migration, migrations idempotentes (Alembic), test en recette d'abord."),
]
r = 5
for rid, desc, imp, prob, mit in risks:
    rc = risk.cell(row=r, column=1, value=rid)
    rc.font = Font(bold=True, size=10, color=WHITE)
    rc.fill = fill(DARK)
    rc.alignment = Alignment(horizontal="center", vertical="center")
    risk.cell(row=r, column=2, value=desc).font = Font(size=10)
    risk.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ic = risk.cell(row=r, column=3, value=imp)
    ic.fill = fill(lvl_fill(imp))
    ic.alignment = Alignment(horizontal="center", vertical="center")
    ic.font = Font(bold=True, size=10)
    pc = risk.cell(row=r, column=4, value=prob)
    pc.fill = fill(lvl_fill(prob))
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.font = Font(bold=True, size=10)
    risk.cell(row=r, column=5, value=mit).font = Font(size=9, color=DARK)
    risk.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in range(1, 6):
        risk.cell(row=r, column=c).border = BORDER
    risk.row_dimensions[r].height = 40
    r += 1

# ==========================================================================
# 9. EXTENSION V2 — EFFICACITÉ DU MODÈLE
# ==========================================================================
eff = wb.create_sheet("8. Efficacité modèle")
eff.sheet_view.showGridLines = False
set_widths(eff, {"A": 7, "B": 24, "C": 60, "D": 11, "E": 8, "F": 14})
title_block(
    eff,
    "Extension V2 — Efficacité du modèle",
    "Qualité + latence : entraînement RunPod (GPU 24 Go) → service K3s CX53 (CPU, GGUF llama.cpp)",
    span=6,
)

# --- Diagnostic (état mesuré au 24 juin 2026) ---
diag = [
    ("Modèle", "Ministral-3-8B-Instruct-2512 · QLoRA r=16/α=32 NF4 · 1 epoch · GGUF Q4_K_M (~5 Go)."),
    ("Service", "llama.cpp CPU sur CX53 (16 vCPU/32 Go), -c 4096, -t 12 ≈ 10-15 tok/s ≈ 1 min/réponse."),
    ("Corpus", "10 000 paires AUTO-générées par Ministral-8B lui-même (pas un maître) + 17 refus + 20 amorce."),
    ("Point faible #1", "Auto-distillation : le fine-tune ne peut PAS dépasser le modèle de base. Plafond de qualité bridé."),
    ("Point faible #2", "Seulement 17 exemples de refus ; éval = 20 cas ; pas de validation pilotant le choix de checkpoint."),
    ("Point faible #3", "Latence ~1 min/réponse (UX) ; quantization Q4_K_M sans imatrix ; pas de cache de préfixe ni KV-quant."),
    ("Atouts en place", "Pipeline RunPod complet, maître souverain dispo (GLM-5.2 / Qwen2.5-72B), RAG actif, journal 👍/👎."),
]
r = 4
for cle, txt in diag:
    cc = eff.cell(row=r, column=1, value=cle)
    cc.fill = fill(DARK if not cle.startswith("Point faible") else ORANGE_DARK)
    cc.font = Font(bold=True, color=WHITE, size=9)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    eff.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
    eff.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    dc = eff.cell(row=r, column=2, value=txt)
    dc.fill = fill(ORANGE_PALE if r % 2 == 0 else WHITE)
    dc.font = Font(size=9, color=DARK)
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in (1, 2):
        eff.cell(row=r, column=c).border = BORDER
    eff.row_dimensions[r].height = 24
    r += 1

# --- Backlog F — leviers d'efficacité ---
r += 1
header_row(eff, r, ["ID", "Lot", "Levier (user story)", "Priorité", "SP", "Phase"])
backlog_eff = [
    ("F1", "1 Mesure", "Étendre l'éval (20 → ~150 cas, multi-tours/refus) + bench de latence + juge GLM-5.2 intégré.", "Must", 5, "Ext-S1"),
    ("F2", "2 Corpus maître", "Distiller depuis un VRAI maître (GLM-5.2 ceiling + Qwen2.5-72B-AWQ souverain) au lieu de l'auto-génération.", "Must", 8, "Ext-S1"),
    ("F3", "2 Corpus maître", "Élargir les refus 17 → ~300 (évasion dose multi-tours, médical, image, hors-filière).", "Must", 5, "Ext-S1"),
    ("F4", "3 Entraînement", "Recette pilotée par l'éval : sweep epochs (1/2/3) × rang LoRA (16/32/64) × lr ; meilleur checkpoint ; max_seq 1024→1536.", "Should", 5, "Ext-S2"),
    ("F5", "4 Alignement", "DPO/ORPO sur ~1-2k paires de préférence (concision « SMS », citation de source, refus nets).", "Should", 8, "Ext-S2"),
    ("F11", "7 Boucle", "Juge GLM-5.2 sur le journal de prod (👍/👎) → corpus curé → ré-entraînement périodique.", "Must", 5, "Ext-S2"),
    ("F6", "5 Quantization", "Quantization imatrix (calibration sur le corpus domaine) → Q4_K_M-imatrix ; comparer Q4_K_M-imatrix vs Q5_K_M.", "Should", 3, "Ext-S3"),
    ("F7", "5 Serving", "llama.cpp : flash-attn, KV-cache q8, cache de préfixe (prompt système figé), threads, max_tokens 512→384.", "Should", 5, "Ext-S3"),
    ("F8", "5 Serving", "Décodage spéculatif (petit modèle brouillon) pour accélérer la génération sur CPU.", "Could", 5, "Ext-S3"),
    ("F9", "6 RAG", "Embeddings plus forts (multilingual-e5 / BGE-M3) + top_k 8 → reranking → 3 ; calibrer le seuil.", "Should", 5, "Ext-S3"),
    ("F10", "6 RAG", "Récupération hybride BM25 + dense (recall sur noms de maladies/variétés).", "Could", 3, "Ext-S3"),
    ("F12", "1 Mesure", "Harnais A/B (qualité + latence) comparant les versions de modèle, choix de la version à déployer.", "Could", 3, "Ext-S3"),
]
PHASE_COLOR = {"Ext-S1": "1565C0", "Ext-S2": "2E7D32", "Ext-S3": "6A1B9A"}
r += 1
for fid, lot, story, prio, sp, phase in backlog_eff:
    eff.cell(row=r, column=1, value=fid).font = Font(bold=True, size=9)
    eff.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    eff.cell(row=r, column=2, value=lot).font = Font(size=8, color=GREY_HEAD)
    eff.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    eff.cell(row=r, column=3, value=story).font = Font(size=10, color=DARK)
    eff.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc = eff.cell(row=r, column=4, value=prio)
    pbg, pfg = MOSCOW[prio]
    pc.fill = fill(pbg)
    pc.font = Font(bold=True, color=pfg, size=10)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    spc = eff.cell(row=r, column=5, value=sp)
    spc.font = Font(bold=True, size=10)
    spc.alignment = Alignment(horizontal="center", vertical="center")
    ph = eff.cell(row=r, column=6, value=phase)
    ph.fill = fill(PHASE_COLOR[phase])
    ph.font = Font(bold=True, color=WHITE, size=9)
    ph.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 7):
        eff.cell(row=r, column=c).border = BORDER
    eff.row_dimensions[r].height = 30
    r += 1

# --- Combinaison recommandée ---
r += 1
eff.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
titre_combi = eff.cell(row=r, column=1, value="Combinaison recommandée (qualité × latence × souveraineté)")
titre_combi.fill = fill(ORANGE)
titre_combi.font = Font(bold=True, color=WHITE, size=11)
titre_combi.alignment = Alignment(horizontal="left", vertical="center", indent=1)
eff.row_dimensions[r].height = 22
r += 1
combi = (
    "Maître : GLM-5.2 (plafond de qualité) sur un sous-ensemble curé + Qwen2.5-72B-AWQ (volume souverain) "
    "→ corpus ~10-12k + 300 refus.  Entraînement : LoRA r=32, 2 epochs, choix de checkpoint par l'éval, "
    "puis DPO sur préférences.  Quantization : Q4_K_M avec imatrix domaine.  Service : flash-attn + KV q8 + "
    "cache de préfixe + max_tokens 384 (threads calibrés).  RAG : embeddings multilingual-e5 + reranking.  "
    "Boucle : juge GLM-5.2 sur le journal → curation → ré-entraînement mensuel.  "
    "Toute combinaison est jugée sur l'éval étendu (F1) : on ne déploie une version que si garde-fous = 100 % "
    "ET qualité ≥ baseline ET latence acceptable."
)
eff.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cc = eff.cell(row=r, column=1, value=combi)
cc.fill = fill(ORANGE_LIGHT)
cc.font = Font(size=10, color=DARK)
cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
eff.row_dimensions[r].height = 92

# Mise à jour de la page de garde : périmètre étendu.
cover["A16"] = "Assistant conversationnel — mémoire, sessions & efficacité du modèle"

wb.save(OUT)
print(f"OK -> {OUT}")
