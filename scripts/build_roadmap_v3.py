"""Génère le backlog agile OpenCacao V3 (suite cacao) au format Excel.

V3 transforme OpenCacao (assistant agronomique) en une SUITE métier complète :
LLM cacao multi-tenant (OpenCacao Pro), intelligence économique du cacao,
traçabilité & conformité EUDR, et modèles prédictifs. Mêmes exigences que la
roadmap V2 (scripts/build_roadmap_v2.py) : page de garde, vision, état des lieux,
backlog MoSCoW/SP, planning Scrum, Gantt, risques, modèle économique.

Usage : python scripts/build_roadmap_v3.py
Sortie : docs/OpenCacao_V3_Backlog_Agile.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# Palette marque OpenLab (orange + anthracite) — identique à la V2.
ORANGE = "EA5B13"
ORANGE_DARK = "B8410A"
ORANGE_LIGHT = "FCE7D8"
ORANGE_PALE = "FDF4EE"
DARK = "1F1F1F"
GREY_HEAD = "3F3F3F"
WHITE = "FFFFFF"

MOSCOW = {
    "Must": ("E53935", WHITE),
    "Should": ("FB8C00", WHITE),
    "Could": ("FDD835", DARK),
    "Won't": ("9E9E9E", WHITE),
}
STATUT = {"À faire": ("E4E4E4", DARK), "En cours": ("BBDEFB", DARK), "Fait": ("C8E6C9", DARK)}
SPRINT_COLOR = {
    "Sprint 1": "EA5B13", "Sprint 2": "F3833E", "Sprint 3": "1565C0", "Sprint 4": "2E7D32",
    "Sprint 5": "6A1B9A", "Sprint 6": "00838F", "Sprint 7": "AD1457", "Sprint 8": "455A64",
}

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROOT = Path(__file__).resolve().parent.parent
LOGO = ROOT / "docs" / "OPENLAB.png"
OUT = ROOT / "docs" / "OpenCacao_V3_Backlog_Agile.xlsx"


def fill(hexcolor: str) -> PatternFill:
    return PatternFill("solid", fgColor=hexcolor)


def header_row(ws, row, headers, *, bg=ORANGE, fg=WHITE, height=26):
    ws.row_dimensions[row].height = height
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill = fill(bg)
        cell.font = Font(bold=True, color=fg, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, subtitle=None, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1, value=title)
    t.fill = fill(DARK)
    t.font = Font(bold=True, color=WHITE, size=16)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.fill = fill(ORANGE)
        s.font = Font(color=WHITE, size=10, italic=True)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20


def zebra_table(ws, start_row, rows, col_key=2, col_val=3, val_span=None):
    """Rend une liste (clé, valeur) en table zébrée à 2 colonnes."""
    r = start_row
    for i, (k, v) in enumerate(rows):
        zc = ORANGE_PALE if i % 2 == 0 else WHITE
        kc = ws.cell(row=r, column=col_key, value=k)
        kc.fill = fill(DARK)
        kc.font = Font(bold=True, color=WHITE, size=10)
        kc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        if val_span:
            ws.merge_cells(start_row=r, start_column=col_val, end_row=r, end_column=val_span)
        vc = ws.cell(row=r, column=col_val, value=v)
        vc.fill = fill(zc)
        vc.font = Font(size=10, color=DARK)
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        kc.border = BORDER
        vc.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1
    return r


wb = Workbook()

# ========================================================================== #
# 1. PAGE DE GARDE                                                           #
# ========================================================================== #
cover = wb.active
cover.title = "Page de garde"
cover.sheet_view.showGridLines = False
for col in "ABCDEFGHIJ":
    cover.column_dimensions[col].width = 13

cover.merge_cells("A1:J1")
cover["A1"].fill = fill(ORANGE)
cover.row_dimensions[1].height = 10

with PILImage.open(LOGO) as im:
    w, h = im.size
target_w = 360
img = XLImage(str(LOGO))
img.width = target_w
img.height = int(h * target_w / w)
img.anchor = "D3"
cover.add_image(img)
for r in range(2, 13):
    cover.row_dimensions[r].height = 22

cover.merge_cells("A15:J15")
cover["A15"] = "OpenCacao — Version 3"
cover["A15"].font = Font(bold=True, size=28, color=DARK)
cover["A15"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[15].height = 44

cover.merge_cells("A16:J16")
cover["A16"] = "De l'assistant à la suite métier : LLM cacao multi-tenant, intelligence économique, conformité EUDR & prédictif"
cover["A16"].font = Font(size=12, color=ORANGE_DARK, italic=True)
cover["A16"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[16].height = 26

cover.merge_cells("A17:J17")
cover["A17"] = "Backlog produit — Gestion de projet Agile (Scrum)"
cover["A17"].font = Font(size=12, color=GREY_HEAD)
cover["A17"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[17].height = 22

cover.merge_cells("C19:H19")
cover["C19"].fill = fill(ORANGE)
cover.row_dimensions[19].height = 4

meta = [
    ("Maître d'ouvrage", "OpenLab Consulting — projet R&D"),
    ("Sponsor / Validation", "Waopron Coulibaly"),
    ("Produit", "OpenCacao — suite IA souveraine pour la filière cacao (Côte d'Ivoire)"),
    ("Périmètre V3", "SaaS multi-tenant (Pro), intelligence économique, traçabilité & conformité EUDR, modèles prédictifs"),
    ("Socle hérité", "V1 (assistant agronomique GGUF/CPU) + V2 (conversationnel, sessions, mémoire)"),
    ("Méthodologie", "Scrum — 8 sprints de 2 semaines (≈ 16 semaines)"),
    ("Date de lancement", "6 octobre 2026"),
    ("Mise en production cible", "25 janvier 2027"),
    ("Version du document", "1.0 — 25 juin 2026"),
    ("Confidentialité", "Interne — R&D OpenLab Consulting"),
]
start = 21
for i, (k, v) in enumerate(meta):
    r = start + i
    cover.row_dimensions[r].height = 22
    cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    kc = cover.cell(row=r, column=2, value=k)
    kc.fill = fill(DARK)
    kc.font = Font(bold=True, color=WHITE, size=10)
    kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    vc = cover.cell(row=r, column=4, value=v)
    vc.fill = fill(ORANGE_PALE if i % 2 == 0 else WHITE)
    vc.font = Font(color=DARK, size=10)
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

foot = start + len(meta) + 1
cover.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=10)
fc = cover.cell(row=foot, column=1, value="OpenLab Consulting — IA souveraine pour la Côte d'Ivoire")
fc.fill = fill(ORANGE)
fc.font = Font(bold=True, color=WHITE, size=10)
fc.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[foot].height = 22

# ========================================================================== #
# 2. VISION & POSITIONNEMENT                                                 #
# ========================================================================== #
vis = wb.create_sheet("1. Vision & positionnement")
vis.sheet_view.showGridLines = False
set_widths(vis, {"A": 3, "B": 30, "C": 96})
title_block(vis, "Vision V3 — d'un assistant à une suite métier cacao",
            "Capitaliser sur l'avantage concurrentiel déjà acquis : la donnée métier + un LLM cacao souverain", span=3)
vision_rows = [
    ("Problème", "La filière cacao ivoirienne (1er producteur mondial) manque d'outils numériques métier : conseil agronomique, "
                 "intelligence de marché et surtout traçabilité conforme au règlement européen anti-déforestation (EUDR), "
                 "désormais condition d'accès au marché UE."),
    ("Proposition de valeur", "Une suite SaaS souveraine bâtie sur le LLM cacao d'OpenCacao : conseil (Pro), intelligence "
                              "économique, conformité EUDR clé en main et modèles prédictifs — en français et langues locales, "
                              "fonctionnant même en zone à faible connectivité."),
    ("Cibles", "Coopératives & unions, exportateurs/négociants, transformateurs, Conseil du Café-Cacao & institutions "
               "(ANADER, FIRCA), banques/assurances agricoles, acheteurs internationaux soumis à l'EUDR."),
    ("Avantage défendable", "Données métier propriétaires + LLM cacao affiné + souveraineté (hébergement local, sans API tierce "
                            "en production) + ancrage terrain. Barrière à l'entrée élevée, concurrence limitée."),
    ("Modèle de revenus", "Abonnement SaaS par siège/coopérative, modules à la carte (Intelligence, EUDR, Prédictif), "
                          "facturation à l'usage (API), prestations d'intégration et de conformité."),
    ("Souveraineté", "Aligné sur le livre blanc « IA souveraine pour la Côte d'Ivoire » : modèle ouvert auto-hébergé, "
                     "données résidentes, conformité à la loi ivoirienne sur les données personnelles."),
]
zebra_table(vis, 4, vision_rows, col_key=2, col_val=3)

# ========================================================================== #
# 3. ÉTAT DES LIEUX                                                          #
# ========================================================================== #
etat = wb.create_sheet("2. État des lieux")
etat.sheet_view.showGridLines = False
set_widths(etat, {"A": 3, "B": 34, "C": 16, "D": 70})
title_block(etat, "Point de départ V3 — l'acquis V1 + V2", "Ce sur quoi la V3 capitalise (au 25 juin 2026)", span=4)
header_row(etat, 4, ["", "Composant", "Statut", "Détail"], bg=GREY_HEAD)
rows = [
    ("LLM cacao souverain", "Acquis", "Ministral-3-8B + LoRA, servi en GGUF/CPU ; garde-fous métier ANADER, sources obligatoires."),
    ("Conseil conversationnel", "Acquis", "V2 : sessions persistantes, mémoire, recherche, identité par appareil, RGPD by design."),
    ("RAG documentaire", "Acquis", "Index dense sur sources officielles (CNRA, ANADER, CCC, FAO, FIRCA) + console de curation."),
    ("Pipeline d'entraînement", "Acquis", "RunPod GPU 24 Go : distillation souveraine (F2), recette pilotée par l'éval (F4), export GGUF."),
    ("Déploiement & CD", "Acquis", "K3s/Hetzner CX53, TLS, livraison continue GitOps (GHCR → ArgoCD)."),
    ("Multi-tenant / comptes org", "À construire", "Pas d'organisations, de rôles ni d'isolation de données par client."),
    ("Facturation / quotas", "À construire", "Aucune offre SaaS, ni mesure d'usage, ni facturation."),
    ("Canaux terrain", "À construire", "Pas de WhatsApp/SMS/USSD ni de mode hors-ligne pour l'agent de terrain."),
    ("Intelligence économique", "À construire", "Pas d'ingestion prix/marché/météo ni de tableaux de bord décisionnels."),
    ("Traçabilité & EUDR", "À construire", "Pas de cartographie de parcelles, de due diligence ni d'attestation déforestation."),
    ("Modèles prédictifs", "À construire", "Pas de prévision de rendement/prix ni de détection de maladies par image."),
]
r = 5
for comp, st, det_txt in rows:
    zc = ORANGE_PALE if (r % 2 == 1) else WHITE
    cc = etat.cell(row=r, column=2, value=comp)
    cc.font = Font(bold=True, size=10, color=DARK)
    cc.fill = fill(zc)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc = etat.cell(row=r, column=3, value=st)
    sc.fill = fill("C8E6C9" if st == "Acquis" else "FFCDD2")
    sc.font = Font(bold=True, size=10, color=DARK)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    dc = etat.cell(row=r, column=4, value=det_txt)
    dc.font = Font(size=10, color=DARK)
    dc.fill = fill(zc)
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in (2, 3, 4):
        etat.cell(row=r, column=c).border = BORDER
    etat.row_dimensions[r].height = 30
    r += 1
concl = r + 1
etat.merge_cells(start_row=concl, start_column=2, end_row=concl, end_column=4)
cc = etat.cell(row=concl, column=2,
               value="Conclusion : le socle technique (LLM cacao + conversationnel + MLOps souverain) est en place. "
                     "La V3 ajoute la couche PRODUIT (SaaS multi-tenant, intelligence, conformité EUDR, prédictif) "
                     "qui transforme la démonstration en offre commerciale.")
cc.fill = fill(ORANGE_LIGHT)
cc.font = Font(bold=True, size=10, color=ORANGE_DARK)
cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
etat.row_dimensions[concl].height = 44

# ========================================================================== #
# 4. BACKLOG PRODUIT V3                                                      #
# ========================================================================== #
backlog = wb.create_sheet("3. Backlog produit")
backlog.sheet_view.showGridLines = False
set_widths(backlog, {"A": 8, "B": 26, "C": 56, "D": 11, "E": 9, "F": 11})
title_block(backlog, "Backlog produit V3 — Epics & User Stories",
            "Priorisation MoSCoW · estimation en Story Points (SP)", span=6)
header_row(backlog, 4, ["ID", "Epic", "User Story", "Priorité", "SP", "Sprint"])

backlog_rows = [
    ("G — Plateforme SaaS multi-tenant", None, None, None, None),
    ("G1", "G — SaaS multi-tenant", "Modèle Organisation/Membre + isolation des données par tenant.", "Must", 8, "Sprint 1"),
    ("G2", "G — SaaS multi-tenant", "Authentification & rôles (RBAC : admin coop, agent, lecteur).", "Must", 5, "Sprint 1"),
    ("G3", "G — SaaS multi-tenant", "Console d'administration d'organisation (membres, paramètres).", "Must", 5, "Sprint 2"),
    ("G4", "G — SaaS multi-tenant", "Mesure d'usage + quotas + facturation par offre/siège.", "Should", 8, "Sprint 2"),
    ("H — Canaux terrain & accessibilité", None, None, None, None),
    ("H1", "H — Canaux terrain", "Connecteur WhatsApp Business (conseil cacao par messagerie).", "Must", 8, "Sprint 3"),
    ("H2", "H — Canaux terrain", "Passerelle SMS/USSD pour zones à faible connectivité.", "Should", 5, "Sprint 3"),
    ("H3", "H — Canaux terrain", "PWA agent de terrain avec mode hors-ligne (sync différée).", "Should", 8, "Sprint 3"),
    ("I — Intelligence économique cacao", None, None, None, None),
    ("I1", "I — Intelligence économique", "Ingestion prix/marché/change + météo + bourses (ICE).", "Must", 8, "Sprint 4"),
    ("I2", "I — Intelligence économique", "Tableaux de bord décisionnels (prix, primes, alertes).", "Must", 5, "Sprint 4"),
    ("I3", "I — Intelligence économique", "Synthèses IA de marché en langage naturel (LLM ancré).", "Should", 5, "Sprint 4"),
    ("J — Traçabilité & conformité EUDR", None, None, None, None),
    ("J1", "J — Traçabilité & EUDR", "Cartographie géolocalisée des parcelles (GPS/polygones).", "Must", 8, "Sprint 6"),
    ("J2", "J — Traçabilité & EUDR", "Croisement cartes déforestation (alerte risque parcelle).", "Must", 8, "Sprint 6"),
    ("J3", "J — Traçabilité & EUDR", "Due diligence & dossier de conformité EUDR exportable.", "Must", 8, "Sprint 7"),
    ("J4", "J — Traçabilité & EUDR", "Traçabilité de lot (coopérative → exportateur → acheteur).", "Should", 5, "Sprint 7"),
    ("K — Modèles prédictifs", None, None, None, None),
    ("K1", "K — Modèles prédictifs", "Prévision de rendement par zone (météo + historique).", "Should", 8, "Sprint 5"),
    ("K2", "K — Modèles prédictifs", "Prévision de prix/saisonnalité (aide à la commercialisation).", "Could", 5, "Sprint 5"),
    ("K3", "K — Modèles prédictifs", "Détection de maladies par image (swollen shoot, pourriture).", "Should", 8, "Sprint 8"),
    ("L — Données, MLOps & souveraineté", None, None, None, None),
    ("L1", "L — Données & MLOps", "Plateforme d'ingestion + gouvernance des données (lignage).", "Must", 5, "Sprint 4"),
    ("L2", "L — Données & MLOps", "Boucle de ré-entraînement continue (journal → curation).", "Should", 5, "Sprint 8"),
    ("L3", "L — Données & MLOps", "Conformité loi ivoirienne données personnelles + RGPD.", "Must", 3, "Sprint 8"),
    ("M — Industrialisation & go-to-market", None, None, None, None),
    ("M1", "M — Industrialisation", "Observabilité, SLA, sauvegardes, plan de reprise (DR).", "Must", 5, "Sprint 8"),
    ("M2", "M — Industrialisation", "Offres commerciales + onboarding coopérative pilote.", "Must", 5, "Sprint 2"),
    ("M3", "M — Industrialisation", "Documentation, recette utilisateur & validation (Waopron).", "Must", 3, "Sprint 8"),
]
r = 5
for row in backlog_rows:
    if row[1] is None:
        backlog.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        gc = backlog.cell(row=r, column=1, value=row[0])
        gc.fill = fill(DARK)
        gc.font = Font(bold=True, color=WHITE, size=11)
        gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        backlog.row_dimensions[r].height = 22
        r += 1
        continue
    rid, epic, story, prio, sp, sprint = row
    backlog.cell(row=r, column=1, value=rid).font = Font(bold=True, size=10)
    backlog.cell(row=r, column=2, value=epic.split("—")[0].strip()).font = Font(size=9, color=GREY_HEAD)
    backlog.cell(row=r, column=3, value=story).font = Font(size=10, color=DARK)
    pc = backlog.cell(row=r, column=4, value=prio)
    pbg, pfg = MOSCOW[prio]
    pc.fill = fill(pbg)
    pc.font = Font(bold=True, color=pfg, size=10)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    spc = backlog.cell(row=r, column=5, value=sp)
    spc.alignment = Alignment(horizontal="center", vertical="center")
    spc.font = Font(bold=True, size=10)
    sc = backlog.cell(row=r, column=6, value=sprint)
    sc.fill = fill(SPRINT_COLOR[sprint])
    sc.font = Font(bold=True, color=WHITE, size=9)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    backlog.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    backlog.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    backlog.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in range(1, 7):
        backlog.cell(row=r, column=c).border = BORDER
    backlog.row_dimensions[r].height = 26
    r += 1
total = r
backlog.merge_cells(start_row=total, start_column=1, end_row=total, end_column=4)
backlog.cell(row=total, column=1, value="TOTAL").font = Font(bold=True, size=11, color=WHITE)
backlog.cell(row=total, column=1).fill = fill(ORANGE)
backlog.cell(row=total, column=1).alignment = Alignment(horizontal="right", vertical="center")
total_sp = sum(x[4] for x in backlog_rows if x[1] is not None)
tc = backlog.cell(row=total, column=5, value=total_sp)
tc.fill = fill(ORANGE)
tc.font = Font(bold=True, size=11, color=WHITE)
tc.alignment = Alignment(horizontal="center", vertical="center")
backlog.cell(row=total, column=6, value="≈18/sprint").fill = fill(ORANGE)
backlog.cell(row=total, column=6).font = Font(bold=True, size=9, color=WHITE)
backlog.cell(row=total, column=6).alignment = Alignment(horizontal="center", vertical="center")

# ========================================================================== #
# 5. PLANNING DES SPRINTS                                                    #
# ========================================================================== #
spr = wb.create_sheet("4. Planning des sprints")
spr.sheet_view.showGridLines = False
set_widths(spr, {"A": 11, "B": 24, "C": 60, "D": 16, "E": 12})
title_block(spr, "Planning des sprints V3", "Scrum — 8 itérations de 2 semaines · cérémonies à chaque sprint", span=5)
header_row(spr, 4, ["Sprint", "Période", "Objectif (incrément livrable)", "User Stories", "Capacité (SP)"])
sprints = [
    ("Sprint 1", "06/10 → 17/10", "Fondations multi-tenant : organisations, RBAC, isolation des données.", "G1, G2", 13),
    ("Sprint 2", "20/10 → 31/10", "Console admin + usage/quotas/facturation + offres & pilote.", "G3, G4, M2", 18),
    ("Sprint 3", "03/11 → 14/11", "Canaux terrain : WhatsApp, SMS/USSD, PWA hors-ligne.", "H1, H2, H3", 21),
    ("Sprint 4", "17/11 → 28/11", "Intelligence économique : ingestion, dashboards, synthèses + données/MLOps.", "I1, I2, I3, L1", 23),
    ("Sprint 5", "01/12 → 12/12", "Prédictif : prévision de rendement & de prix.", "K1, K2", 13),
    ("Sprint 6", "15/12 → 09/01", "Traçabilité : cartographie parcelles + croisement déforestation.", "J1, J2", 16),
    ("Sprint 7", "12/01 → 16/01", "Conformité EUDR : due diligence, dossier exportable, traçabilité de lot.", "J3, J4", 13),
    ("Sprint 8", "19/01 → 25/01", "Prédictif image + MLOps + durcissement, conformité, recette.", "K3, L2, L3, M1, M3", 21),
]
r = 5
for s, per, obj, us, cap in sprints:
    sc = spr.cell(row=r, column=1, value=s)
    sc.fill = fill(SPRINT_COLOR[s])
    sc.font = Font(bold=True, color=WHITE, size=11)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=2, value=per).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=2).font = Font(bold=True, size=10)
    spr.cell(row=r, column=3, value=obj).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    spr.cell(row=r, column=3).font = Font(size=10)
    spr.cell(row=r, column=4, value=us).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=4).font = Font(size=10)
    spr.cell(row=r, column=5, value=cap).alignment = Alignment(horizontal="center", vertical="center")
    spr.cell(row=r, column=5).font = Font(bold=True, size=10)
    for c in range(1, 6):
        spr.cell(row=r, column=c).border = BORDER
    spr.row_dimensions[r].height = 38
    r += 1
note_r = r + 1
spr.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=5)
nc = spr.cell(row=note_r, column=1,
              value="Définition de « Terminé » (héritée V2) : code revu, testé (≥ 80 %), garde-fous métier verts, "
                    "isolation multi-tenant vérifiée, documenté, déployé en recette. Le sprint 6 intègre les congés de fin d'année.")
nc.fill = fill(ORANGE_PALE)
nc.font = Font(italic=True, size=10, color=DARK)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
spr.row_dimensions[note_r].height = 40

# ========================================================================== #
# 6. GANTT                                                                   #
# ========================================================================== #
gantt = wb.create_sheet("5. Gantt")
gantt.sheet_view.showGridLines = False
weeks = ["S1\n06/10", "S2\n13/10", "S3\n20/10", "S4\n27/10", "S5\n03/11", "S6\n10/11",
         "S7\n17/11", "S8\n24/11", "S9\n01/12", "S10\n08/12", "S11\n15/12", "S12\n05/01",
         "S13\n12/01", "S14\n19/01"]
set_widths(gantt, {"A": 8, "B": 40})
for i in range(len(weeks)):
    gantt.column_dimensions[get_column_letter(3 + i)].width = 6
title_block(gantt, "Diagramme de Gantt — vue hebdomadaire", "≈ 16 semaines · 8 sprints de 2 semaines", span=2 + len(weeks))
gantt.row_dimensions[4].height = 30
gantt.cell(row=4, column=1, value="ID").fill = fill(GREY_HEAD)
gantt.cell(row=4, column=2, value="User Story").fill = fill(GREY_HEAD)
for c in (1, 2):
    gantt.cell(row=4, column=c).font = Font(bold=True, color=WHITE, size=10)
    gantt.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center")
    gantt.cell(row=4, column=c).border = BORDER
for i, wlabel in enumerate(weeks):
    cell = gantt.cell(row=4, column=3 + i, value=wlabel)
    cell.fill = fill(GREY_HEAD)
    cell.font = Font(bold=True, color=WHITE, size=8)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
sprint_weeks = {
    "Sprint 1": (0, 1), "Sprint 2": (2, 3), "Sprint 3": (4, 5), "Sprint 4": (6, 7),
    "Sprint 5": (8, 9), "Sprint 6": (10, 11), "Sprint 7": (12, 12), "Sprint 8": (13, 13),
}
gantt_rows = [
    ("G1", "Multi-tenant + isolation", "Sprint 1"),
    ("G2", "RBAC rôles", "Sprint 1"),
    ("G3", "Console admin org", "Sprint 2"),
    ("G4", "Usage / quotas / facturation", "Sprint 2"),
    ("M2", "Offres + pilote coopérative", "Sprint 2"),
    ("H1", "WhatsApp Business", "Sprint 3"),
    ("H2", "SMS / USSD", "Sprint 3"),
    ("H3", "PWA hors-ligne", "Sprint 3"),
    ("I1", "Ingestion prix/marché/météo", "Sprint 4"),
    ("I2", "Tableaux de bord", "Sprint 4"),
    ("I3", "Synthèses IA de marché", "Sprint 4"),
    ("L1", "Données + gouvernance", "Sprint 4"),
    ("K1", "Prévision rendement", "Sprint 5"),
    ("K2", "Prévision prix", "Sprint 5"),
    ("J1", "Cartographie parcelles", "Sprint 6"),
    ("J2", "Risque déforestation", "Sprint 6"),
    ("J3", "Dossier EUDR exportable", "Sprint 7"),
    ("J4", "Traçabilité de lot", "Sprint 7"),
    ("K3", "Détection maladies (image)", "Sprint 8"),
    ("L2", "Ré-entraînement continu", "Sprint 8"),
    ("L3", "Conformité données", "Sprint 8"),
    ("M1", "Observabilité / SLA / DR", "Sprint 8"),
    ("M3", "Doc + recette finale", "Sprint 8"),
]
r = 5
for rid, label, sprint in gantt_rows:
    gantt.cell(row=r, column=1, value=rid).font = Font(bold=True, size=9)
    gantt.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    gantt.cell(row=r, column=2, value=label).font = Font(size=9)
    gantt.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    gantt.cell(row=r, column=1).border = BORDER
    gantt.cell(row=r, column=2).border = BORDER
    w0, w1 = sprint_weeks[sprint]
    for i in range(len(weeks)):
        cell = gantt.cell(row=r, column=3 + i)
        cell.border = BORDER
        if w0 <= i <= w1:
            cell.fill = fill(SPRINT_COLOR[sprint])
    gantt.row_dimensions[r].height = 18
    r += 1
gantt.freeze_panes = "C5"

# ========================================================================== #
# 7. RISQUES                                                                 #
# ========================================================================== #
risk = wb.create_sheet("6. Risques")
risk.sheet_view.showGridLines = False
set_widths(risk, {"A": 6, "B": 46, "C": 12, "D": 12, "E": 56})
title_block(risk, "Registre des risques V3", "Suivi en rétrospective de sprint", span=5)
header_row(risk, 4, ["ID", "Risque", "Impact", "Probabilité", "Mitigation"])

def lvl_fill(v):
    return {"Élevé": "FFCDD2", "Moyen": "FFE0B2", "Faible": "C8E6C9",
            "Élevée": "FFCDD2", "Moyenne": "FFE0B2"}.get(v, "FFE0B2")

risks = [
    ("R1", "Évolution / calendrier du règlement EUDR (report, modalités) impactant la valeur du module conformité.", "Élevé", "Moyenne", "Veille réglementaire ; architecture modulaire ; valeur traçabilité utile même hors EUDR."),
    ("R2", "Qualité/fraîcheur des données de marché et cartographiques (déforestation, prix).", "Élevé", "Moyenne", "Sources multiples redondantes, contrôles qualité, mention d'incertitude, partenariats données."),
    ("R3", "Isolation multi-tenant insuffisante (fuite de données entre coopératives).", "Élevé", "Faible", "Isolation par tenant testée, revues sécurité, tests d'intrusion avant go-live."),
    ("R4", "Adoption terrain freinée par la connectivité et la littératie numérique.", "Élevé", "Moyenne", "Canaux SMS/USSD/WhatsApp + mode hors-ligne + onboarding accompagné des coopératives."),
    ("R5", "Capacité d'infrastructure (CPU CX53) insuffisante pour la charge SaaS multi-tenant.", "Élevé", "Moyenne", "Dimensionnement progressif, mise à l'échelle horizontale, file d'inférence, monitoring."),
    ("R6", "Dépendance à un acteur institutionnel (Conseil du Café-Cacao) pour les données officielles.", "Moyen", "Moyenne", "Conventions de données, sources alternatives, valeur produit indépendante des données officielles."),
    ("R7", "Modèle prédictif peu fiable au lancement (données historiques limitées).", "Moyen", "Élevée", "Démarrer en aide à la décision (intervalles de confiance), améliorer par la boucle de données."),
    ("R8", "Conformité à la loi ivoirienne sur les données personnelles / hébergement.", "Élevé", "Faible", "Données résidentes, minimisation, consentement, registre de traitement dès la conception."),
]
r = 5
for rid, desc, imp, prob, mit in risks:
    rc = risk.cell(row=r, column=1, value=rid)
    rc.font = Font(bold=True, size=10, color=WHITE)
    rc.fill = fill(DARK)
    rc.alignment = Alignment(horizontal="center", vertical="center")
    risk.cell(row=r, column=2, value=desc).font = Font(size=10)
    risk.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ic = risk.cell(row=r, column=3, value=imp)
    ic.fill = fill(lvl_fill(imp))
    ic.alignment = Alignment(horizontal="center", vertical="center")
    ic.font = Font(bold=True, size=10)
    pc = risk.cell(row=r, column=4, value=prob)
    pc.fill = fill(lvl_fill(prob))
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.font = Font(bold=True, size=10)
    risk.cell(row=r, column=5, value=mit).font = Font(size=9, color=DARK)
    risk.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in range(1, 6):
        risk.cell(row=r, column=c).border = BORDER
    risk.row_dimensions[r].height = 40
    r += 1

# ========================================================================== #
# 8. MODÈLE ÉCONOMIQUE                                                       #
# ========================================================================== #
biz = wb.create_sheet("7. Modèle économique")
biz.sheet_view.showGridLines = False
set_widths(biz, {"A": 3, "B": 26, "C": 40, "D": 54})
title_block(biz, "Modèle économique V3", "Offre SaaS modulaire — souveraine, à la coopérative et à l'usage", span=4)
header_row(biz, 4, ["", "Offre", "Cible & modules", "Logique de prix"])
offres = [
    ("Conseil (Pro)", "Coopératives, agents ANADER — conseil agronomique + conversationnel", "Abonnement par siège/agent ; palier coopérative."),
    ("Intelligence", "Exportateurs, négociants, institutions — prix, marché, météo, alertes", "Abonnement module + options data premium."),
    ("Conformité EUDR", "Exportateurs & acheteurs UE — cartographie, due diligence, dossier", "Abonnement + à la parcelle/au lot tracé."),
    ("Prédictif", "Coopératives, banques/assurances — rendement, prix, maladies", "Module + facturation à l'usage (API/scoring)."),
    ("Intégration", "Toutes cibles — déploiement souverain, formation, support", "Prestation projet + maintenance/SLA."),
]
r = 5
for nom, cible, prix in offres:
    zc = ORANGE_PALE if r % 2 == 1 else WHITE
    nc = biz.cell(row=r, column=2, value=nom)
    nc.fill = fill(DARK)
    nc.font = Font(bold=True, color=WHITE, size=10)
    nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cc = biz.cell(row=r, column=3, value=cible)
    cc.fill = fill(zc)
    cc.font = Font(size=10, color=DARK)
    cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc = biz.cell(row=r, column=4, value=prix)
    pc.fill = fill(zc)
    pc.font = Font(size=10, color=DARK)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for c in (2, 3, 4):
        biz.cell(row=r, column=c).border = BORDER
    biz.row_dimensions[r].height = 34
    r += 1
note = r + 1
biz.merge_cells(start_row=note, start_column=2, end_row=note, end_column=4)
nc = biz.cell(row=note, column=2,
              value="Trajectoire : coopérative pilote (preuve de valeur) → déploiement régional via unions/Conseil du Café-Cacao → "
                    "module EUDR comme produit d'appel pour les exportateurs soumis au marché UE. Détail marché & chiffrage : "
                    "business plan OpenLab Consulting.")
nc.fill = fill(ORANGE_LIGHT)
nc.font = Font(italic=True, size=10, color=ORANGE_DARK)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
biz.row_dimensions[note].height = 46

wb.save(OUT)
print(f"OK -> {OUT}")
